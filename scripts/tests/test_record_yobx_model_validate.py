"""Tests for ``scripts.record_yobx_model_validate``."""

from __future__ import annotations

import json
import os
import sys
import tempfile
import unittest

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.dirname(HERE))

import record_yobx_model_validate as rymv  # noqa: E402


class TestRecordYobxModelValidate(unittest.TestCase):
    def test_defaults(self):
        model_ids = {e["model"] for e in rymv.DEFAULT_MODELS}
        self.assertIn("arnir0/Tiny-LLM", model_ids)
        self.assertIn("mistralai/Mistral-7B-v0.3", model_ids)
        # Each model entry must declare the per-model fields used by the
        # recorder so the snapshot can carry them.
        for entry in rymv.DEFAULT_MODELS:
            self.assertIn("model", entry)
            self.assertIn("dtype", entry)
            self.assertIn("device", entry)
            self.assertIn("atol", entry)
            self.assertIn("task", entry)
        labels = {e["label"] for e in rymv.DEFAULT_EXPORTERS}
        self.assertIn("yobx", labels)
        self.assertIn("yobx-ort", labels)
        self.assertIn("dynamo-ir", labels)
        self.assertIn("onnx-dynamo-os_ort", labels)
        self.assertIn("yobx-to_onnx", labels)
        # Each exporter config must declare the three required fields.
        for cfg in rymv.DEFAULT_EXPORTERS:
            self.assertIn("label", cfg)
            self.assertIn("exporter", cfg)
            self.assertIn("optimization", cfg)
        self.assertEqual(rymv.DEFAULT_DTYPE, "float16")
        self.assertEqual(rymv.DEFAULT_DEVICE, "cpu")

    def test_coerce_model_entry_from_string(self):
        entry = rymv._coerce_model_entry("a/b")
        self.assertEqual(entry["model"], "a/b")
        self.assertEqual(entry["dtype"], rymv.DEFAULT_DTYPE)
        self.assertEqual(entry["device"], rymv.DEFAULT_DEVICE)
        self.assertEqual(entry["atol"], rymv.DEFAULT_ATOL)
        self.assertEqual(entry["task"], rymv.DEFAULT_TASK)

    def test_coerce_model_entry_from_dict_preserves_overrides(self):
        src = dict(model="a/b", dtype="float32", atol=0.5, device="cuda", task="t")
        entry = rymv._coerce_model_entry(src)
        self.assertEqual(entry["model"], "a/b")
        self.assertEqual(entry["dtype"], "float32")
        self.assertEqual(entry["device"], "cuda")
        self.assertEqual(entry["atol"], 0.5)
        self.assertEqual(entry["task"], "t")
        # Returns a copy so the input is not mutated.
        self.assertIsNot(entry, src)

    def test_stringify_error_collapses_whitespace_and_truncates(self):
        self.assertEqual(rymv._stringify_error(None), "")
        self.assertEqual(rymv._stringify_error("boom"), "boom")
        # Multi-line errors are collapsed into a single line so that the
        # actionable trailing lines (such as HuggingFace's "You need to
        # have sentencepiece installed ...") are preserved in the snapshot.
        self.assertEqual(rymv._stringify_error("boom\nrest"), "boom rest")
        self.assertEqual(
            rymv._stringify_error(
                "Couldn't instantiate the backend tokenizer from one of: \n"
                "(1) a `tokenizers` library serialization file, \n"
                "(2) a slow tokenizer instance to convert or \n"
                "(3) an equivalent slow tokenizer class to instantiate and convert. \n"
                "You need to have sentencepiece installed to convert a slow "
                "tokenizer to a fast one."
            ),
            "Couldn't instantiate the backend tokenizer from one of: "
            "(1) a `tokenizers` library serialization file, "
            "(2) a slow tokenizer instance to convert or "
            "(3) an equivalent slow tokenizer class to instantiate and convert. "
            "You need to have sentencepiece installed to convert a slow "
            "tokenizer to a fast one.",
        )
        long = "x" * 500
        out = rymv._stringify_error(long)
        self.assertTrue(out.endswith("..."))
        self.assertEqual(len(out), 400)

    def test_to_float_handles_non_finite(self):
        self.assertIsNone(rymv._to_float(None))
        self.assertIsNone(rymv._to_float("nope"))
        self.assertIsNone(rymv._to_float(float("inf")))
        self.assertIsNone(rymv._to_float(float("nan")))
        self.assertEqual(rymv._to_float(1.5), 1.5)
        self.assertEqual(rymv._to_float("2"), 2.0)

    def test_is_cell_working(self):
        self.assertTrue(rymv.is_cell_working({"export": "OK", "discrepancies": "OK"}))
        self.assertFalse(
            rymv.is_cell_working({"export": "FAILED", "discrepancies": "OK"})
        )
        self.assertFalse(
            rymv.is_cell_working({"export": "OK", "discrepancies": "FAILED"})
        )
        # Export OK but no discrepancy check -> not "working" for our purposes.
        self.assertFalse(rymv.is_cell_working({"export": "OK"}))
        self.assertFalse(rymv.is_cell_working({}))
        self.assertFalse(rymv.is_cell_working(None))

    def test_is_cell_working_honours_model_atol(self):
        # Even when ``validate_model`` reports a failed discrepancy check
        # (because it used a stricter default tolerance), a cell where the
        # observed max abs error is within the model's declared atol should
        # be considered working.
        summary = {
            "export": "OK",
            "discrepancies": "FAILED",
            "discrepancies_max_abs": 0.01,
            "discrepancies_atol": 1e-5,
        }
        self.assertTrue(rymv.is_cell_working(summary, model_atol=0.02))
        # Still failing when the per-model atol is stricter than the error.
        self.assertFalse(rymv.is_cell_working(summary, model_atol=0.001))
        # Export failure dominates even when within atol.
        self.assertFalse(
            rymv.is_cell_working({**summary, "export": "FAILED"}, model_atol=0.02)
        )
        # Without a max_abs we cannot conclude success.
        self.assertFalse(
            rymv.is_cell_working(
                {"export": "OK", "discrepancies": "FAILED"}, model_atol=0.02
            )
        )

    def test_first_error_picks_earliest_failing_step(self):
        step, msg = rymv._first_error(
            {
                "error_config": "bad config\nstack",
                "error_export": "export boom",
            }
        )
        self.assertEqual(step, "config")
        self.assertEqual(msg, "bad config stack")

        step, msg = rymv._first_error({"error_export": "export boom"})
        self.assertEqual(step, "export")
        self.assertEqual(msg, "export boom")

        step, msg = rymv._first_error({})
        self.assertEqual(step, "")
        self.assertEqual(msg, "")

    def test_normalise_result_success(self):
        cfg = {"label": "yobx", "exporter": "yobx", "optimization": "default"}
        row = rymv._normalise_result(
            "arnir0/Tiny-LLM",
            cfg,
            {
                "export": "OK",
                "discrepancies": "OK",
                "discrepancies_ok": 3,
                "discrepancies_total": 3,
                "discrepancies_max_abs": 1e-5,
                "discrepancies_atol": 1e-3,
                "n_nodes": 42,
                "top_op_types": "MatMul:5",
            },
            duration_s=1.25,
        )
        self.assertEqual(row["model_id"], "arnir0/Tiny-LLM")
        self.assertEqual(row["label"], "yobx")
        self.assertEqual(row["exporter"], "yobx")
        self.assertEqual(row["optimization"], "default")
        self.assertEqual(row["success"], 1)
        self.assertEqual(row["export"], "OK")
        self.assertEqual(row["discrepancies"], "OK")
        self.assertEqual(row["discrepancies_ok"], 3)
        self.assertEqual(row["discrepancies_total"], 3)
        self.assertAlmostEqual(row["discrepancies_max_abs"], 1e-5)
        self.assertAlmostEqual(row["discrepancies_atol"], 1e-3)
        self.assertEqual(row["n_nodes"], 42)
        self.assertEqual(row["top_op_types"], "MatMul:5")
        self.assertAlmostEqual(row["duration_s"], 1.25)
        self.assertEqual(row["error_step"], "")
        self.assertEqual(row["error"], "")

    def test_normalise_result_failure(self):
        cfg = {"label": "dynamo-ir", "exporter": "dynamo", "optimization": "ir"}
        row = rymv._normalise_result(
            "microsoft/Phi-4-reasoning",
            cfg,
            {"export": "FAILED", "error_export": "boom!\nstack"},
        )
        self.assertEqual(row["success"], 0)
        self.assertEqual(row["error_step"], "export")
        self.assertEqual(row["error"], "boom! stack")
        self.assertIsNone(row["discrepancies_max_abs"])
        self.assertEqual(row["discrepancies"], "")

    def test_normalise_result_success_within_model_atol(self):
        # ``validate_model`` flagged discrepancies as failed (because of its
        # stricter default tolerance), but the observed max abs error is
        # below the per-model atol, so the cell should be marked as working.
        cfg = {"label": "yobx", "exporter": "yobx", "optimization": "default"}
        row = rymv._normalise_result(
            "arnir0/Tiny-LLM",
            cfg,
            {
                "export": "OK",
                "discrepancies": "FAILED",
                "discrepancies_max_abs": 0.01,
                "discrepancies_atol": 1e-5,
            },
            model_atol=0.02,
        )
        self.assertEqual(row["success"], 1)
        self.assertEqual(row["error_step"], "")
        # ``discrepancies`` and ``discrepancies_atol`` should be rewritten so
        # the cached JSON reflects the per-model tolerance that made the cell
        # successful, instead of the stricter default used by ``validate_model``.
        self.assertEqual(row["discrepancies"], "OK")
        self.assertAlmostEqual(row["discrepancies_atol"], 0.02)
        self.assertEqual(row["error"], "")

    def test_merge_last_working_records_now_on_success(self):
        row = {"success": 1}
        out = rymv.merge_last_working(row, None, "2024-05-01T00:00:00Z", "abc")
        self.assertEqual(out["last_working_date"], "2024-05-01T00:00:00Z")
        self.assertEqual(out["last_working_commit"], "abc")

    def test_merge_last_working_preserves_previous_on_failure(self):
        row = {"success": 0}
        previous = {
            "last_working_date": "2024-01-01T00:00:00Z",
            "last_working_commit": "deadbeef",
        }
        out = rymv.merge_last_working(row, previous, "2024-05-01T00:00:00Z", "abc")
        self.assertEqual(out["last_working_date"], "2024-01-01T00:00:00Z")
        self.assertEqual(out["last_working_commit"], "deadbeef")

    def test_merge_last_working_no_previous_failure_yields_empty(self):
        row = {"success": 0}
        out = rymv.merge_last_working(row, None, "2024-05-01T00:00:00Z", "abc")
        self.assertEqual(out["last_working_date"], "")
        self.assertEqual(out["last_working_commit"], "")

    def test_build_payload_totals_and_dates(self):
        cfg_yobx = {"label": "yobx", "exporter": "yobx", "optimization": "default"}
        cfg_dyn = {"label": "dynamo-ir", "exporter": "dynamo", "optimization": "ir"}
        raw = [
            ("arnir0/Tiny-LLM", cfg_yobx, {"export": "OK", "discrepancies": "OK"}, 1.5),
            (
                "arnir0/Tiny-LLM",
                cfg_dyn,
                {"export": "FAILED", "error_export": "boom"},
                0.7,
            ),
            (
                "microsoft/Phi-4-reasoning",
                cfg_yobx,
                {"export": "FAILED", "error_export": "boom"},
                2.0,
            ),
            (
                "microsoft/Phi-4-reasoning",
                cfg_dyn,
                {"export": "OK", "discrepancies": "OK"},
                3.25,
            ),
        ]
        previous = {
            "results": [
                {
                    "model_id": "microsoft/Phi-4-reasoning",
                    "label": "yobx",
                    "last_working_date": "2024-01-01T00:00:00Z",
                    "last_working_commit": "old",
                }
            ],
            "tasks": {"microsoft/Phi-4-reasoning": "fill-mask"},
        }
        payload = rymv.build_payload(
            raw_results=raw,
            models=(
                dict(
                    model="arnir0/Tiny-LLM",
                    dtype="float16",
                    atol=0.02,
                    device="cpu",
                    task="text-generation",
                ),
                dict(
                    model="microsoft/Phi-4-reasoning",
                    dtype="float16",
                    atol=0.02,
                    device="cpu",
                    task="text-generation",
                ),
            ),
            exporters=(cfg_yobx, cfg_dyn),
            dtype="float16",
            device="cpu",
            commit="newcommit",
            previous_payload=previous,
            now="2024-05-01T00:00:00Z",
            tasks={"arnir0/Tiny-LLM": "text-generation"},
        )
        self.assertEqual(payload["date"], "2024-05-01T00:00:00Z")
        self.assertEqual(payload["dtype"], "float16")
        self.assertEqual(payload["device"], "cpu")
        self.assertEqual(
            payload["totals"]["yobx"], {"success": 1, "failure": 1, "total": 2}
        )
        self.assertEqual(
            payload["totals"]["dynamo-ir"], {"success": 1, "failure": 1, "total": 2}
        )
        rows_by_key = {(r["model_id"], r["label"]): r for r in payload["results"]}
        # New success refreshes last_working.
        self.assertEqual(
            rows_by_key[("arnir0/Tiny-LLM", "yobx")]["last_working_date"],
            "2024-05-01T00:00:00Z",
        )
        # Previous-known good is preserved on a fresh failure.
        self.assertEqual(
            rows_by_key[("microsoft/Phi-4-reasoning", "yobx")]["last_working_date"],
            "2024-01-01T00:00:00Z",
        )
        # No previous and now failing -> empty.
        self.assertEqual(
            rows_by_key[("arnir0/Tiny-LLM", "dynamo-ir")]["last_working_date"], ""
        )
        # Per-cell export duration is preserved in the row.
        self.assertAlmostEqual(
            rows_by_key[("arnir0/Tiny-LLM", "yobx")]["duration_s"], 1.5
        )
        self.assertAlmostEqual(
            rows_by_key[("microsoft/Phi-4-reasoning", "dynamo-ir")]["duration_s"], 3.25
        )
        # HuggingFace task is captured per row, taking the explicit value first
        # and falling back to the previous snapshot.
        self.assertEqual(
            rows_by_key[("arnir0/Tiny-LLM", "yobx")]["task"], "text-generation"
        )
        self.assertEqual(
            rows_by_key[("microsoft/Phi-4-reasoning", "yobx")]["task"],
            "text-generation",
        )
        self.assertEqual(
            payload["tasks"],
            {
                "arnir0/Tiny-LLM": "text-generation",
                "microsoft/Phi-4-reasoning": "text-generation",
            },
        )
        # Per-model dtype/device/atol are recorded in each row.
        self.assertEqual(rows_by_key[("arnir0/Tiny-LLM", "yobx")]["dtype"], "float16")
        self.assertEqual(rows_by_key[("arnir0/Tiny-LLM", "yobx")]["device"], "cpu")
        self.assertAlmostEqual(rows_by_key[("arnir0/Tiny-LLM", "yobx")]["atol"], 0.02)
        # Payload must be JSON-serialisable with allow_nan=False.
        json.dumps(payload, allow_nan=False)

    def test_load_existing_cache_missing_or_invalid(self):
        with tempfile.TemporaryDirectory() as tmp:
            self.assertEqual(rymv._load_existing_cache(os.path.join(tmp, "x.json")), {})
            bad = os.path.join(tmp, "bad.json")
            with open(bad, "w", encoding="utf-8") as fh:
                fh.write("not json")
            self.assertEqual(rymv._load_existing_cache(bad), {})
            arr = os.path.join(tmp, "arr.json")
            with open(arr, "w", encoding="utf-8") as fh:
                fh.write("[]")
            self.assertEqual(rymv._load_existing_cache(arr), {})

    def test_parse_args_defaults(self):
        args = rymv.parse_args([])
        self.assertEqual(args.cache_dir, os.path.join("cache_data"))
        self.assertEqual(args.repo, "yet-another-onnx-builder")
        self.assertIsNone(args.models)
        self.assertEqual(args.dtype, "float16")
        self.assertEqual(args.device, "cpu")
        self.assertIsNone(args.limit)
        self.assertIsNone(args.dump_folder)

    def test_parse_args_dump_folder(self):
        args = rymv.parse_args(["--dump-folder", "/tmp/dump"])
        self.assertEqual(args.dump_folder, "/tmp/dump")

    def test_parse_args_quiet_default(self):
        args = rymv.parse_args([])
        self.assertTrue(args.quiet)

    def test_parse_args_no_quiet(self):
        args = rymv.parse_args(["--no-quiet"])
        self.assertFalse(args.quiet)
        args = rymv.parse_args(["--no-quiet", "--quiet"])
        self.assertTrue(args.quiet)

    def test_parse_args_custom_models(self):
        args = rymv.parse_args(["--model", "a/b", "--model", "c/d", "--limit", "1"])
        self.assertEqual(args.models, ["a/b", "c/d"])
        self.assertEqual(args.limit, 1)

    def _write_extra_xlsx(self, path: str, value: float) -> None:
        """Write a minimal workbook with the ``extra`` sheet used by yobx."""
        from openpyxl import Workbook

        wb = Workbook()
        # Replace the default sheet with one named ``extra`` and populate it
        # with a small key/value table mimicking the yobx export report.
        default = wb.active
        wb.remove(default)
        ws = wb.create_sheet("extra")
        ws.append(["name", "value"])
        ws.append(["builder", "torch"])
        ws.append(["stat_time_export_and_post_processing", value])
        ws.append(["stat_time_post_process_exported_program", 0.001])
        wb.save(path)

    def test_read_yobx_export_duration_returns_metric(self):
        try:
            import openpyxl  # noqa: F401
        except Exception:
            self.skipTest("openpyxl is not installed")
        with tempfile.TemporaryDirectory() as tmp:
            xlsx = os.path.join(tmp, "model.xlsx")
            self._write_extra_xlsx(xlsx, 1.5)
            self.assertAlmostEqual(rymv._read_yobx_export_duration(xlsx), 1.5)

    def test_read_yobx_export_duration_missing_file(self):
        self.assertIsNone(rymv._read_yobx_export_duration("/no/such/file.xlsx"))

    def test_read_yobx_export_duration_missing_sheet(self):
        try:
            from openpyxl import Workbook
        except Exception:
            self.skipTest("openpyxl is not installed")
        with tempfile.TemporaryDirectory() as tmp:
            xlsx = os.path.join(tmp, "other.xlsx")
            wb = Workbook()
            wb.active.append(["a", "b"])
            wb.save(xlsx)
            self.assertIsNone(rymv._read_yobx_export_duration(xlsx))

    def test_list_xlsx_recurses(self):
        with tempfile.TemporaryDirectory() as tmp:
            self.assertEqual(rymv._list_xlsx(tmp), set())
            sub = os.path.join(tmp, "sub")
            os.makedirs(sub)
            a = os.path.join(tmp, "a.xlsx")
            b = os.path.join(sub, "b.xlsx")
            with open(a, "w", encoding="utf-8") as fh:
                fh.write("")
            with open(b, "w", encoding="utf-8") as fh:
                fh.write("")
            # A non-xlsx file is ignored.
            with open(os.path.join(tmp, "skip.txt"), "w", encoding="utf-8") as fh:
                fh.write("")
            self.assertEqual(rymv._list_xlsx(tmp), {a, b})
            self.assertEqual(rymv._list_xlsx(None), set())
            self.assertEqual(rymv._list_xlsx("/no/such/folder"), set())

    def test_run_all_uses_xlsx_duration_for_yobx(self):
        try:
            import openpyxl  # noqa: F401
        except Exception:
            self.skipTest("openpyxl is not installed")
        cfg_yobx = {"label": "yobx", "exporter": "yobx", "optimization": "default"}
        cfg_dyn = {"label": "dynamo-ir", "exporter": "dynamo", "optimization": "ir"}

        with tempfile.TemporaryDirectory() as tmp:
            test = self

            def fake_run_validate_one(
                entry, exporter_cfg, verbose=0, dump_folder=None, quiet=True
            ):
                if exporter_cfg["exporter"] == "yobx":
                    # Mimic the yobx exporter writing a companion workbook.
                    xlsx = os.path.join(
                        dump_folder, f"{entry['model'].replace('/', '_')}.xlsx"
                    )
                    test._write_extra_xlsx(xlsx, 2.5)
                return {"export": "OK", "discrepancies": "OK"}

            original = rymv.run_validate_one
            rymv.run_validate_one = fake_run_validate_one
            try:
                results = rymv.run_all(
                    models=({"model": "a/b", "dtype": "float16", "device": "cpu"},),
                    exporters=(cfg_yobx, cfg_dyn),
                    dump_folder=tmp,
                )
            finally:
                rymv.run_validate_one = original

        durations = {item[1]["label"]: item[3] for item in results}
        # yobx duration is replaced by the metric from the extra sheet.
        self.assertAlmostEqual(durations["yobx"], 2.5)
        # The other exporter keeps its wall-clock measurement.
        self.assertGreaterEqual(durations["dynamo-ir"], 0.0)
        self.assertNotAlmostEqual(durations["dynamo-ir"], 2.5)

    def test_run_all_uses_temp_dump_folder_when_none_provided(self):
        """When no dump folder is supplied, yobx still gets one (a temp dir)."""
        try:
            import openpyxl  # noqa: F401
        except Exception:
            self.skipTest("openpyxl is not installed")
        cfg_yobx = {"label": "yobx", "exporter": "yobx", "optimization": "default"}
        test = self
        seen_folders = []

        def fake_run_validate_one(
            entry, exporter_cfg, verbose=0, dump_folder=None, quiet=True
        ):
            # yobx must receive a real dump folder so it can save the report.
            test.assertIsNotNone(dump_folder)
            test.assertTrue(os.path.isdir(dump_folder))
            seen_folders.append(dump_folder)
            xlsx = os.path.join(dump_folder, "model.xlsx")
            test._write_extra_xlsx(xlsx, 3.75)
            return {"export": "OK", "discrepancies": "OK"}

        original = rymv.run_validate_one
        rymv.run_validate_one = fake_run_validate_one
        try:
            results = rymv.run_all(
                models=({"model": "a/b", "dtype": "float16", "device": "cpu"},),
                exporters=(cfg_yobx,),
                dump_folder=None,
            )
        finally:
            rymv.run_validate_one = original

        self.assertAlmostEqual(results[0][3], 3.75)
        # The temporary directory is cleaned up after the metric is read.
        self.assertEqual(len(seen_folders), 1)
        self.assertFalse(os.path.isdir(seen_folders[0]))

    def test_existing_snapshot_is_valid_json(self):
        repo_root = os.path.dirname(os.path.dirname(HERE))
        path = os.path.join(
            repo_root,
            "cache_data",
            "yet-another-onnx-builder",
            "model_validate.json",
        )
        if not os.path.exists(path):
            self.skipTest("model_validate.json snapshot not present in repo")
        with open(path, encoding="utf-8") as fh:

            def _reject(token: str) -> None:
                raise ValueError(f"non-JSON token in snapshot: {token}")

            payload = json.load(fh, parse_constant=_reject)
        for key in (
            "date",
            "exporters",
            "models",
            "results",
            "totals",
            "dtype",
            "device",
        ):
            self.assertIn(key, payload)
        self.assertIsInstance(payload["results"], list)

    def test_is_rate_limit_error_detects_429(self):
        self.assertTrue(
            rymv._is_rate_limit_error(
                Exception(
                    "HTTP Error 429 thrown while requesting HEAD "
                    "https://huggingface.co/mistralai/Mistral-7B-v0.3/"
                    "resolve/main/config.json"
                )
            )
        )
        self.assertTrue(
            rymv._is_rate_limit_error(Exception("429 Client Error: Too Many Requests"))
        )

        class _Resp:
            status_code = 429

        class _Err(Exception):
            response = _Resp()

        self.assertTrue(rymv._is_rate_limit_error(_Err("boom")))

    def test_is_rate_limit_error_ignores_other_errors(self):
        self.assertFalse(rymv._is_rate_limit_error(Exception("HTTP Error 404")))
        self.assertFalse(rymv._is_rate_limit_error(ValueError("bad model")))

    def test_is_hf_hub_access_error_detects_offline_message(self):
        msg = (
            "We couldn't connect to 'https://huggingface.co' to load the "
            "files, and couldn't find them in the cached files. Check "
            "your internet connection or see how to run the library in "
            "offline mode at "
            "'https://huggingface.co/docs/transformers/installation"
            "#offline-mode'."
        )
        self.assertTrue(rymv._is_hf_hub_access_error(msg))
        self.assertTrue(rymv._is_hf_hub_access_error(Exception(msg)))

    def test_is_hf_hub_access_error_ignores_unrelated_errors(self):
        self.assertFalse(rymv._is_hf_hub_access_error(None))
        self.assertFalse(rymv._is_hf_hub_access_error(""))
        self.assertFalse(rymv._is_hf_hub_access_error("ValueError: bad model"))
        # A rate-limit error mentions huggingface.co but is not a Hub
        # access error in the sense detected here.
        self.assertFalse(
            rymv._is_hf_hub_access_error(
                "HTTP Error 429 thrown while requesting HEAD "
                "https://huggingface.co/x/y/resolve/main/config.json"
            )
        )

    def test_run_all_short_circuits_subsequent_exporters_on_hub_access_error(self):
        cfgs = (
            {"label": "yobx", "exporter": "yobx", "optimization": "default"},
            {"label": "dynamo-ir", "exporter": "dynamo", "optimization": "ir"},
        )
        offline_msg = (
            "We couldn't connect to 'https://huggingface.co' to load the "
            "files, and couldn't find them in the cached files. Check "
            "your internet connection or see how to run the library in "
            "offline mode."
        )
        calls = []

        def fake(entry, exporter_cfg, verbose=0, dump_folder=None, quiet=True):
            calls.append((entry["model"], exporter_cfg["label"]))
            return {
                "model_id": entry["model"],
                "export": "FAILED",
                "error_config": offline_msg,
            }

        original = rymv.run_validate_one
        rymv.run_validate_one = fake
        try:
            out = rymv.run_all(
                models=(
                    {"model": "gated/model", "dtype": "float16", "device": "cpu"},
                    {"model": "ok/model", "dtype": "float16", "device": "cpu"},
                ),
                exporters=cfgs,
                dump_folder=None,
                quiet=True,
            )
        finally:
            rymv.run_validate_one = original

        # Only the first exporter is executed for ``gated/model`` (the
        # second is short-circuited), then both exporters run for the next
        # model even though it also fails (each model is short-circuited
        # independently).
        self.assertEqual(
            calls,
            [
                ("gated/model", "yobx"),
                ("ok/model", "yobx"),
            ],
        )
        # All four (model, exporter) cells are still present in the
        # returned list so the dashboard payload keeps a full grid.
        self.assertEqual(len(out), 4)
        labels = [(model_id, cfg["label"]) for model_id, cfg, _, _ in out]
        self.assertEqual(
            labels,
            [
                ("gated/model", "yobx"),
                ("gated/model", "dynamo-ir"),
                ("ok/model", "yobx"),
                ("ok/model", "dynamo-ir"),
            ],
        )
        # The short-circuited cell carries the same error_config message
        # so the dashboard still surfaces the root cause.
        skipped_summary = out[1][2]
        self.assertEqual(skipped_summary["export"], "FAILED")
        self.assertEqual(skipped_summary["error_config"], offline_msg)
        self.assertEqual(out[1][3], 0.0)

    def test_run_all_aborts_on_http_429(self):
        cfg = {"label": "yobx", "exporter": "yobx", "optimization": "default"}

        def boom(entry, exporter_cfg, verbose=0, dump_folder=None, quiet=True):
            raise RuntimeError(
                "HTTP Error 429 thrown while requesting HEAD "
                "https://huggingface.co/mistralai/Mistral-7B-v0.3/"
                "resolve/main/config.json"
            )

        original = rymv.run_validate_one
        rymv.run_validate_one = boom
        try:
            with self.assertRaises(RuntimeError) as ctx:
                rymv.run_all(
                    models=({"model": "a/b", "dtype": "float16", "device": "cpu"},),
                    exporters=(cfg,),
                    dump_folder=None,
                    quiet=True,
                )
            self.assertIn("429", str(ctx.exception))
        finally:
            rymv.run_validate_one = original

    def test_run_validate_one_dispatches_to_to_onnx_default(self):
        """``yobx-to_onnx`` exporter routes to ``run_to_onnx_default``."""
        cfg = {
            "label": "yobx-to_onnx",
            "exporter": "yobx-to_onnx",
            "optimization": "(defaults)",
        }
        seen = {}

        def fake_to_onnx_default(
            entry, exporter_cfg, verbose=0, dump_folder=None, quiet=True
        ):
            seen["entry"] = entry
            seen["exporter_cfg"] = exporter_cfg
            seen["dump_folder"] = dump_folder
            seen["quiet"] = quiet
            return {"export": "OK", "discrepancies": "OK"}

        original = rymv.run_to_onnx_default
        rymv.run_to_onnx_default = fake_to_onnx_default
        try:
            result = rymv.run_validate_one(
                {"model": "a/b", "dtype": "float16", "device": "cpu"},
                cfg,
                dump_folder="/tmp/x",
                quiet=True,
            )
        finally:
            rymv.run_to_onnx_default = original
        self.assertEqual(result, {"export": "OK", "discrepancies": "OK"})
        self.assertEqual(seen["entry"]["model"], "a/b")
        self.assertEqual(seen["exporter_cfg"]["exporter"], "yobx-to_onnx")
        self.assertEqual(seen["dump_folder"], "/tmp/x")
        self.assertTrue(seen["quiet"])


    def test_run_validate_one_dispatches_to_olive_modelbuilder(self):
        """``olive-modelbuilder`` exporter routes to ``run_olive_modelbuilder``."""
        cfg = {
            "label": "olive-modelbuilder",
            "exporter": "olive-modelbuilder",
            "optimization": "(modelbuilder)",
        }
        seen = {}

        def fake_olive_modelbuilder(
            entry, exporter_cfg, verbose=0, dump_folder=None, quiet=True
        ):
            seen["entry"] = entry
            seen["exporter_cfg"] = exporter_cfg
            seen["dump_folder"] = dump_folder
            seen["quiet"] = quiet
            return {"export": "OK", "discrepancies": "SKIPPED"}

        original = rymv.run_olive_modelbuilder
        rymv.run_olive_modelbuilder = fake_olive_modelbuilder
        try:
            result = rymv.run_validate_one(
                {"model": "a/b", "dtype": "float16", "device": "cpu"},
                cfg,
                dump_folder="/tmp/x",
                quiet=True,
            )
        finally:
            rymv.run_olive_modelbuilder = original
        self.assertEqual(result, {"export": "OK", "discrepancies": "SKIPPED"})
        self.assertEqual(seen["entry"]["model"], "a/b")
        self.assertEqual(seen["exporter_cfg"]["exporter"], "olive-modelbuilder")
        self.assertEqual(seen["dump_folder"], "/tmp/x")
        self.assertTrue(seen["quiet"])

    def test_is_cell_working_treats_skipped_discrepancies_as_ok(self):
        """``SKIPPED`` discrepancies (e.g. ``olive-modelbuilder``) are working."""
        self.assertTrue(
            rymv.is_cell_working({"export": "OK", "discrepancies": "SKIPPED"})
        )
        self.assertFalse(
            rymv.is_cell_working({"export": "FAILED", "discrepancies": "SKIPPED"})
        )

    def test_olive_precision_for_dtype(self):
        self.assertEqual(rymv._olive_precision_for_dtype("float16"), "fp16")
        self.assertEqual(rymv._olive_precision_for_dtype("fp16"), "fp16")
        self.assertEqual(rymv._olive_precision_for_dtype("bfloat16"), "bf16")
        self.assertEqual(rymv._olive_precision_for_dtype("int4"), "int4")
        self.assertEqual(rymv._olive_precision_for_dtype("float32"), "fp32")
        self.assertEqual(rymv._olive_precision_for_dtype(None), "fp32")

    def test_default_exporters_include_olive_modelbuilder(self):
        labels = {e["label"] for e in rymv.DEFAULT_EXPORTERS}
        self.assertIn("olive-modelbuilder", labels)
        # The exporter dispatch key must match the value handled by
        # ``run_validate_one``.
        cfg = next(
            e for e in rymv.DEFAULT_EXPORTERS if e["label"] == "olive-modelbuilder"
        )
        self.assertEqual(cfg["exporter"], "olive-modelbuilder")

    def test_parse_args_test_flag(self):
        args = rymv.parse_args([])
        self.assertFalse(args.test)
        args = rymv.parse_args(["--test"])
        self.assertTrue(args.test)

    def test_run_olive_modelbuilder_when_cli_missing(self):
        """``run_olive_modelbuilder`` returns a failed summary when the CLI is missing."""
        import subprocess

        original_run = subprocess.run

        def fake_run(*args, **kwargs):
            raise FileNotFoundError("olive not installed")

        subprocess.run = fake_run
        try:
            summary = rymv.run_olive_modelbuilder(
                {"model": "a/b", "dtype": "float16", "device": "cpu"},
                {
                    "label": "olive-modelbuilder",
                    "exporter": "olive-modelbuilder",
                    "optimization": "(modelbuilder)",
                },
            )
        finally:
            subprocess.run = original_run
        self.assertEqual(summary["export"], "FAILED")
        self.assertIn("olive", summary["error_export"].lower())

    def test_run_olive_modelbuilder_non_zero_returncode(self):
        """``run_olive_modelbuilder`` surfaces the CLI's stderr on failure."""
        import subprocess

        class _Proc:
            returncode = 2
            stdout = ""
            stderr = "boom: missing model"

        original_run = subprocess.run

        def fake_run(*args, **kwargs):
            return _Proc()

        subprocess.run = fake_run
        try:
            summary = rymv.run_olive_modelbuilder(
                {"model": "a/b", "dtype": "float16", "device": "cpu"},
            {
                "label": "olive-modelbuilder",
                "exporter": "olive-modelbuilder",
                "optimization": "(modelbuilder)",
            },
            )
        finally:
            subprocess.run = original_run
        self.assertEqual(summary["export"], "FAILED")
        self.assertEqual(
            summary["error_export"], "--dry_run failed: boom: missing model"
        )

    def test_run_olive_modelbuilder_reads_discrepancy_check_results(self):
        """A successful run reads metrics from ``discrepancy_check_results.json``."""
        import subprocess

        with tempfile.TemporaryDirectory() as tmp:
            class _Proc:
                returncode = 0
                stdout = ""
                stderr = ""

            captured = {"calls": []}

            def fake_run(cmd, *args, **kwargs):
                # Drop a fake ONNX model and discrepancy results in the
                # ``--output_path`` directory the recorder passes in.
                idx = cmd.index("--output_path")
                output_path = cmd[idx + 1]
                captured["calls"].append(list(cmd))
                captured["cmd"] = list(cmd)
                onnx_path = os.path.join(output_path, "model.onnx")
                with open(onnx_path, "wb") as f:
                    f.write(b"\x00")
                with open(
                    os.path.join(output_path, "discrepancy_check_results.json"),
                    "w",
                    encoding="utf-8",
                ) as f:
                    json.dump(
                        {
                            "max_abs_error": 0.0125,
                            "elements_above_0_1": 0,
                            "elements_above_0_01": 3,
                            "total_elements": 1000,
                            "status": "passed",
                        },
                        f,
                    )
                return _Proc()

            original_run = subprocess.run
            subprocess.run = fake_run
            try:
                summary = rymv.run_olive_modelbuilder(
                    {"model": "a/b", "dtype": "float16", "device": "cpu"},
                    {
                        "label": "olive-modelbuilder",
                        "exporter": "olive-modelbuilder",
                        "optimization": "(modelbuilder)",
                    },
                    dump_folder=tmp,
                )
            finally:
                subprocess.run = original_run

        self.assertEqual(summary["export"], "OK")
        self.assertEqual(summary["discrepancies"], "OK")
        self.assertEqual(summary["discrepancies_total"], 1000)
        self.assertEqual(summary["discrepancies_ok"], 997)
        self.assertAlmostEqual(summary["discrepancies_max_abs"], 0.0125)
        self.assertEqual(summary["discrepancies_atol"], 0.01)
        # ``--test`` must be passed so Olive auto-injects the
        # ``OnnxDiscrepancyCheck`` pass and dumps the JSON metrics file.
        self.assertIn("--test", captured["cmd"])
        # Two commands are issued: a ``--dry_run`` one (to save the
        # workflow ``config.json`` including the ``OnnxDiscrepancyCheck``
        # pass) followed by the actual run. Both must include ``--test``
        # so Olive injects the discrepancy-check pass in both.
        self.assertEqual(len(captured["calls"]), 2)
        self.assertIn("--dry_run", captured["calls"][0])
        self.assertIn("--test", captured["calls"][0])
        self.assertNotIn("--dry_run", captured["calls"][1])
        self.assertIn("--test", captured["calls"][1])

    def test_run_olive_modelbuilder_missing_discrepancy_results(self):
        """Successful export but no JSON metrics file is reported as a discrepancy failure."""
        import subprocess

        with tempfile.TemporaryDirectory() as tmp:
            class _Proc:
                returncode = 0
                stdout = ""
                stderr = ""

            def fake_run(cmd, *args, **kwargs):
                idx = cmd.index("--output_path")
                output_path = cmd[idx + 1]
                with open(os.path.join(output_path, "model.onnx"), "wb") as f:
                    f.write(b"\x00")
                # Intentionally do not write discrepancy_check_results.json
                return _Proc()

            original_run = subprocess.run
            subprocess.run = fake_run
            try:
                summary = rymv.run_olive_modelbuilder(
                    {"model": "a/b", "dtype": "float16", "device": "cpu"},
                    {
                        "label": "olive-modelbuilder",
                        "exporter": "olive-modelbuilder",
                        "optimization": "(modelbuilder)",
                    },
                    dump_folder=tmp,
                )
            finally:
                subprocess.run = original_run

        self.assertEqual(summary["export"], "OK")
        self.assertEqual(summary["discrepancies"], "FAILED")
        self.assertIn(
            "discrepancy_check_results.json", summary["error_discrepancies"]
        )


class TestPerModelPerExporterDispatch(unittest.TestCase):
    """One wiring check per (model, exporter) cell of the model-id coverage page.

    The ``dashboard/yet-another-onnx-builder/model-validate.html`` page is fed
    by one cell per model (``DEFAULT_MODELS``) and per exporter
    (``DEFAULT_EXPORTERS``). This exercises every such cell to make sure
    ``run_validate_one`` routes it to the right backend so that both the Olive
    runtime (``olive-modelbuilder``) and the ``torch.onnx.export`` based
    columns (the ``dynamo``/``onnx-dynamo`` exporters, which ``validate_model``
    drives through ``torch.onnx.export``) are actually invoked. ``yobx`` and
    ``yobx-to_onnx`` are checked too for completeness.
    """

    def _install_fake_validate_model(self, recorder):
        """Inject a fake ``yobx.torch.validate`` so the lazy import resolves.

        ``run_validate_one`` imports ``validate_model`` lazily inside the
        function, so a fake module tree is enough to intercept the call
        without requiring the heavy ``yobx``/``torch`` stack to be installed.
        Returns a restore token to feed back to :meth:`_restore_modules`.
        """
        import types

        names = ("yobx", "yobx.torch", "yobx.torch.validate")
        saved = {name: sys.modules.get(name) for name in names}

        yobx_mod = types.ModuleType("yobx")
        torch_mod = types.ModuleType("yobx.torch")
        validate_mod = types.ModuleType("yobx.torch.validate")

        def fake_validate_model(**kwargs):
            recorder["kwargs"] = kwargs
            return ({"export": "OK", "discrepancies": "OK"}, {"data": 1})

        validate_mod.validate_model = fake_validate_model
        yobx_mod.torch = torch_mod
        torch_mod.validate = validate_mod
        sys.modules["yobx"] = yobx_mod
        sys.modules["yobx.torch"] = torch_mod
        sys.modules["yobx.torch.validate"] = validate_mod
        return saved

    def _restore_modules(self, saved):
        for name, mod in saved.items():
            if mod is None:
                sys.modules.pop(name, None)
            else:
                sys.modules[name] = mod

    def _check_special_cased(self, entry, cfg, attr):
        """Assert ``run_validate_one`` routes the cell to ``rymv.<attr>``."""
        seen = {}

        def fake_backend(entry, exporter_cfg, verbose=0, dump_folder=None, quiet=True):
            seen["entry"] = entry
            seen["exporter_cfg"] = exporter_cfg
            seen["dump_folder"] = dump_folder
            seen["quiet"] = quiet
            return {"export": "OK", "discrepancies": "OK"}

        original = getattr(rymv, attr)
        setattr(rymv, attr, fake_backend)
        try:
            result = rymv.run_validate_one(
                entry, cfg, dump_folder="/tmp/x", quiet=True
            )
        finally:
            setattr(rymv, attr, original)
        self.assertEqual(result, {"export": "OK", "discrepancies": "OK"})
        self.assertEqual(seen["entry"]["model"], entry["model"])
        self.assertEqual(seen["exporter_cfg"]["exporter"], cfg["exporter"])
        self.assertEqual(seen["dump_folder"], "/tmp/x")

    def _check_validate_model(self, entry, cfg):
        """Assert ``run_validate_one`` calls ``validate_model`` for the cell."""
        recorder = {}
        saved = self._install_fake_validate_model(recorder)
        try:
            result = rymv.run_validate_one(
                entry, cfg, dump_folder="/tmp/x", quiet=True
            )
        finally:
            self._restore_modules(saved)
        self.assertEqual(result, {"export": "OK", "discrepancies": "OK"})
        kwargs = recorder["kwargs"]
        self.assertEqual(kwargs["model_id"], entry["model"])
        self.assertEqual(kwargs["exporter"], cfg["exporter"])
        self.assertEqual(kwargs["optimization"], cfg["optimization"])

    def _check_cell(self, entry, cfg):
        exporter = cfg["exporter"]
        if exporter == "olive-modelbuilder":
            self._check_special_cased(entry, cfg, "run_olive_modelbuilder")
        elif exporter == "yobx-to_onnx":
            self._check_special_cased(entry, cfg, "run_to_onnx_default")
        else:
            self._check_validate_model(entry, cfg)


def _slugify(value: str) -> str:
    """Return a valid Python identifier fragment for a test method name."""
    return "".join(ch if ch.isalnum() else "_" for ch in value)


def _make_cell_test(entry, cfg):
    def test(self):
        self._check_cell(entry, cfg)

    test.__doc__ = (
        f"model-id coverage cell ({entry['model']!r}, {cfg['label']!r}) "
        "is routed to the right backend."
    )
    return test


# Generate one test function per (model, exporter) cell of the model-id
# coverage page so that a failure pinpoints the exact cell that regressed
# (and so each model/exporter pairing is its own test rather than a single
# parametrised loop). This covers the Olive runtime (``olive-modelbuilder``)
# and the ``torch.onnx.export`` based columns (``dynamo``/``onnx-dynamo``)
# alongside ``yobx`` and ``yobx-to_onnx``.
for _entry in rymv.DEFAULT_MODELS:
    for _cfg in rymv.DEFAULT_EXPORTERS:
        _name = (
            f"test_cell_{_slugify(_entry['model'])}__{_slugify(_cfg['label'])}"
        )
        setattr(
            TestPerModelPerExporterDispatch,
            _name,
            _make_cell_test(_entry, _cfg),
        )
del _entry, _cfg, _name


if __name__ == "__main__":
    unittest.main()
