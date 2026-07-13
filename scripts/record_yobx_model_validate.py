"""Record ``yobx validate`` discrepancy snapshots for selected HuggingFace models.

This script exercises :func:`yobx.torch.validate.validate_model` (the same
entry point used by ``python -m yobx validate``) over a fixed set of model
ids and a fixed set of ONNX exporter configurations. The resulting summary
for each (model, exporter) cell is serialised to
``cache_data/yet-another-onnx-builder/model_validate.json``.

The JSON file is consumed by
``dashboard/yet-another-onnx-builder/model-validate.html`` which renders a
small table showing whether the model still exports without discrepancies
for each exporter and what was the last date the cell was working.

For every cell, the previous snapshot is consulted so that the
``last_working_date`` (and ``last_working_commit``) of a cell is preserved
when a previously working configuration starts failing. The intent is to
make it obvious when an export regressed and when it was last known good.

The run uses ``dtype=float16`` and ``device=cpu`` and exercises the
following exporter configurations:

* ``yobx`` with ``optimization='default'``
* ``yobx`` with ``optimization='default+onnxruntime'``
* ``dynamo`` with ``optimization='ir'``
* ``onnx-dynamo`` with ``optimization='os_ort'``
* ``olive-modelbuilder`` -- a runtime based on the development version
  of `Olive <https://github.com/microsoft/Olive>`_. It invokes the
  ``olive capture-onnx-graph --use_model_builder`` CLI (the
  ``ModelBuilder`` pass, backed by ``onnxruntime-genai``) and reports
  whether the export succeeded. No discrepancy check is run for this
  column: ``ModelBuilder`` rebuilds the graph from scratch from the
  HuggingFace config so a direct numerical comparison against the
  PyTorch reference is not meaningful in the same way.

Usage::

    python scripts/record_yobx_model_validate.py [--cache-dir DIR]
        [--model ID ...] [--limit N] [--test]

The ``--test`` flag is a smoke-test shortcut that restricts the run to a
single tiny model (``arnir0/Tiny-LLM``). It is intended to quickly
validate that a newly-added runtime (such as the ``olive-modelbuilder``
one) is wired correctly without hitting the full default list of models.
"""

from __future__ import annotations

import argparse
from contextlib import nullcontext
import datetime as dt
import json
import math
import os
import sys
import time
from typing import Any, Dict, List, Optional, Tuple
from unittest.mock import patch

# ``sentencepiece`` (used by SentencePiece-tokenised models such as
# ``mistralai/Mistral-7B-v0.3``) parses its model file with generated
# protobuf bindings that require the pure-Python implementation when the
# C++-backed ``protobuf`` wheel pulled in by ``onnx`` (>=4.25.1) is
# installed.  Forcing the Python implementation here – at module level,
# before any ``sentencepiece`` import – is the upstream-documented fix
# for the ``Couldn't instantiate the backend tokenizer … You need to
# have sentencepiece or tiktoken installed`` error that otherwise shows
# up in the dashboard even when ``sentencepiece`` *is* installed.
os.environ.setdefault("PROTOCOL_BUFFERS_PYTHON_IMPLEMENTATION", "python")

DEFAULT_DTYPE = "float16"
DEFAULT_DEVICE = "cpu"
DEFAULT_ATOL = 0.02
DEFAULT_RTOL = 10
DEFAULT_TASK = "text-generation"


def _is_rate_limit_error(exc: BaseException) -> bool:
    """Return ``True`` when *exc* looks like a HuggingFace HTTP 429 error.

    The HuggingFace Hub raises rate-limit errors as
    ``huggingface_hub.utils.HfHubHTTPError`` (or plain ``requests`` /
    ``urllib`` errors) whose message contains either ``HTTP Error 429`` or
    ``Too Many Requests``. When such an error happens, retrying within the
    same CI run is pointless and only burns rate limiting budget, so the
    caller must abort instead of merely logging the failure.
    """
    message = f"{exc}"
    if "429" in message and (
        "Too Many Requests" in message
        or "HTTP Error 429" in message
        or "rate limit" in message.lower()
    ):
        return True
    status = getattr(getattr(exc, "response", None), "status_code", None)
    return status == 429


def _is_hf_hub_access_error(value: Any) -> bool:
    """Return ``True`` when *value* looks like a HuggingFace Hub access error.

    Both ``transformers`` and ``huggingface_hub`` surface a misleading
    ``We couldn't connect to 'https://huggingface.co' to load the files,
    and couldn't find them in the cached files`` message for *any* failure
    to reach the Hub: a real network outage, a missing/invalid
    ``HF_TOKEN``, or a gated model the current token is not allowed to
    download (for example ``mistralai/Mistral-7B-v0.3``).

    Detecting this pattern lets the recorder short-circuit the remaining
    exporter cells for the same model: they would all fail at the same
    ``error_config`` step with the same long message, producing redundant
    noise in the dashboard and wasting rate-limit budget against the Hub.

    *value* may be either an exception instance or a plain string (the
    summary returned by ``validate_model`` stores the error as a string in
    ``error_config``).
    """
    if value is None:
        return False
    text = str(value)
    lower = text.lower()
    if "huggingface.co" not in lower:
        return False
    return (
        "couldn't connect to" in lower
        or "couldn t connect to" in lower
        or "offline mode" in lower
    )


def default_fp16_tg(model_id):
    return dict(
        model=model_id,
        dtype="float16",
        atol=DEFAULT_ATOL,
        rtol=DEFAULT_RTOL,
        device="cpu",
        task="text-generation",
    )


DEFAULT_MODELS: Tuple[Dict[str, Any], ...] = (
    default_fp16_tg("arnir0/Tiny-LLM"),
    {**default_fp16_tg("mistralai/Mistral-7B-v0.3"), "tokenizer_use_fast": False},
)

# Each exporter configuration is fully described by a small dict so that the
# dashboard can render meaningful column headers. ``label`` is the unique
# identifier used in the JSON snapshot (both in ``exporters`` and in the
# ``label`` field of each result row).
DEFAULT_EXPORTERS: Tuple[Dict[str, str], ...] = (
    {"label": "yobx", "exporter": "yobx", "optimization": "default"},
    {
        "label": "yobx-ort",
        "exporter": "yobx",
        "optimization": "default+onnxruntime",
    },
    {"label": "dynamo-ir", "exporter": "dynamo", "optimization": "ir"},
    {
        "label": "onnx-dynamo-os_ort",
        "exporter": "onnx-dynamo",
        "optimization": "os_ort",
    },
    # Runtime based on the development version of Olive. It calls the
    # ``olive capture-onnx-graph --use_model_builder --test`` CLI which
    # runs the ``ModelBuilder`` pass (backed by ``onnxruntime-genai``)
    # to convert the HuggingFace model to ONNX, then auto-injects the
    # ``OnnxDiscrepancyCheck`` pass to compare the resulting graph
    # against a randomly-initialised two-hidden-layer copy of the same
    # architecture and dumps the metrics to
    # ``discrepancy_check_results.json`` (see :func:`run_olive_modelbuilder`).
    {
        "label": "olive-modelbuilder",
        "exporter": "olive-modelbuilder",
        "optimization": "(modelbuilder)",
    },
)


def _coerce_model_entry(
    item: Any,
    *,
    default_dtype: str = DEFAULT_DTYPE,
    default_device: str = DEFAULT_DEVICE,
    default_atol: float = DEFAULT_ATOL,
    default_task: str = DEFAULT_TASK,
) -> Dict[str, Any]:
    """Return a fully-populated model entry dict from a string id or dict.

    Accepted forms:

    * ``"org/model"`` – a bare HuggingFace model id; defaults are filled in.
    * ``{"model": ..., "dtype": ..., "device": ..., "atol": ..., "task": ...}``
      – any missing key falls back to the supplied default.
    """
    if isinstance(item, str):
        data: Dict[str, Any] = {"model": item}
    elif isinstance(item, dict):
        data = dict(item)
    else:
        raise TypeError(f"Model entry must be a str or dict, got {type(item).__name__}")
    if not data.get("model"):
        raise ValueError(f"Model entry is missing a 'model' field: {item!r}")
    data.setdefault("dtype", default_dtype)
    data.setdefault("device", default_device)
    data.setdefault("atol", default_atol)
    data.setdefault("task", default_task)
    return data


def _log(message: str) -> None:
    """Print ``message`` prefixed with a UTC timestamp."""
    now = dt.datetime.now(tz=dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    print(f"[{now}] {message}", flush=True)


def collect_versions() -> Dict[str, str]:
    """Return the versions of the relevant packages, if importable."""
    versions: Dict[str, str] = {}
    for name in (
        "yobx",
        "torch",
        "transformers",
        "onnx",
        "onnxruntime",
        "onnxscript",
        "olive",
        "onnxruntime_genai",
    ):
        try:
            module = __import__(name)
        except Exception:  # noqa: BLE001 - best effort, optional packages
            continue
        version = getattr(module, "__version__", None)
        if version:
            versions[name] = str(version)
    return versions


def _stringify_error(value: Any) -> str:
    """Return a short, single-line string for an exporter error message.

    Multi-line error messages are collapsed into a single line (newlines and
    runs of whitespace are replaced with a single space) instead of keeping
    only the first line. HuggingFace ``transformers`` in particular raises
    multi-line errors whose *first* line is uninformative on its own (for
    example ``Couldn't instantiate the backend tokenizer from one of:``) and
    whose subsequent lines contain the actual remediation hint (``You need
    to have sentencepiece installed ...``). Keeping the whole message - up
    to a hard cap - makes the JSON snapshot consumed by the dashboard
    actually useful for debugging failing model exports.
    """
    if value is None:
        return ""
    text = " ".join(str(value).split())
    if len(text) > 400:
        text = text[:397] + "..."
    return text


def _to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    try:
        f = float(value)
    except (TypeError, ValueError):
        return None
    # JSON does not support Infinity/NaN; coerce to None so the payload
    # remains valid JSON parsable by browsers.
    if not math.isfinite(f):
        return None
    return f


def _to_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _summary_get(summary: Any, key: str, default: Any = None) -> Any:
    """Best-effort getter that works for both dicts and ``ValidateSummary``."""
    if summary is None:
        return default
    try:
        return summary.get(key, default)
    except AttributeError:
        pass
    if hasattr(summary, key):
        return getattr(summary, key, default)
    try:
        return summary[key]
    except (TypeError, KeyError, IndexError):
        return default


def is_cell_working(summary: Any, model_atol: Optional[float] = None) -> bool:
    """Return ``True`` when the export *and* the discrepancy check succeeded.

    ``summary`` may be either a ``ValidateSummary`` instance or a plain dict
    (the latter is used by the tests).

    When ``model_atol`` is provided, a cell is also considered working if the
    export succeeded and the maximum absolute discrepancy is below the
    per-model tolerance, even if ``validate_model`` flagged the run as failed
    because it used a stricter default tolerance.
    """
    export = _summary_get(summary, "export")
    discrepancies = _summary_get(summary, "discrepancies")
    if export != "OK":
        return False
    if discrepancies == "OK":
        return True
    # ``SKIPPED`` means the runtime intentionally did not run the
    # discrepancy check (for instance the ``olive-modelbuilder`` column,
    # which rebuilds the graph from scratch and cannot be compared
    # numerically against the live torch module). For such cells a
    # successful export is enough to mark the cell as working.
    if discrepancies == "SKIPPED":
        return True
    # ``discrepancies`` was set but reported as failed by ``validate_model``.
    # If the caller supplied a per-model ``atol`` that is more permissive than
    # the one used during validation, honour it: a cell where the observed
    # maximum absolute error is within the model's declared tolerance is still
    # considered working for our dashboard.
    if model_atol is not None and discrepancies:
        max_abs = _to_float(_summary_get(summary, "discrepancies_max_abs"))
        if max_abs is not None and max_abs <= model_atol:
            return True
    # ``discrepancies`` is only set when ``do_run=True``. If it was not set
    # the export ran but we cannot conclude that the model is "working"
    # numerically, so we conservatively report False.
    return False


def _first_error(summary: Any) -> Tuple[str, str]:
    """Return ``(error_step, error_message)`` for the first failing step."""
    steps = (
        ("config", "error_config"),
        ("tokenizer", "error_tokenizer"),
        ("model", "error_model"),
        ("observer", "error_observer"),
        ("export", "error_export"),
        ("discrepancies", "error_discrepancies"),
    )
    for step, field in steps:
        value = _summary_get(summary, field)
        if value:
            return step, _stringify_error(value)
    return "", ""


def _normalise_result(
    model_id: str,
    exporter_cfg: Dict[str, str],
    summary: Any,
    duration_s: Optional[float] = None,
    model_atol: Optional[float] = None,
) -> Dict[str, Any]:
    """Pick a JSON-serialisable subset of the fields returned by ``validate_model``."""
    working = is_cell_working(summary, model_atol=model_atol)
    error_step, error = _first_error(summary)

    discrepancies = _summary_get(summary, "discrepancies") or ""
    discrepancies_atol = _to_float(_summary_get(summary, "discrepancies_atol"))
    # When the cell is considered working only because the observed error is
    # within the per-model ``atol`` (which is more permissive than the default
    # tolerance used by ``validate_model``), override the failure flag and the
    # tolerance reported in the snapshot so the JSON is self-consistent.
    if working and model_atol is not None and discrepancies and discrepancies != "OK":
        discrepancies = "OK"
        discrepancies_atol = float(model_atol)
        if error_step == "discrepancies":
            error_step, error = "", ""

    return {
        "model_id": model_id,
        "label": exporter_cfg["label"],
        "exporter": exporter_cfg["exporter"],
        "optimization": exporter_cfg["optimization"],
        "success": 1 if working else 0,
        "export": _summary_get(summary, "export") or "",
        "discrepancies": discrepancies,
        "discrepancies_ok": _to_int(_summary_get(summary, "discrepancies_ok")),
        "discrepancies_total": _to_int(_summary_get(summary, "discrepancies_total")),
        "discrepancies_max_abs": _to_float(
            _summary_get(summary, "discrepancies_max_abs")
        ),
        "discrepancies_atol": discrepancies_atol,
        "n_nodes": _to_int(_summary_get(summary, "n_nodes")),
        "top_op_types": _summary_get(summary, "top_op_types") or "",
        "duration_s": _to_float(duration_s),
        "error_step": error_step,
        "error": error,
    }


def _load_existing_cache(path: str) -> Dict[str, Any]:
    """Read the previous snapshot if it exists, returning ``{}`` on any error."""
    if not path or not os.path.exists(path):
        return {}
    try:
        with open(path, encoding="utf-8") as fh:
            payload = json.load(fh)
    except (OSError, ValueError):
        return {}
    if not isinstance(payload, dict):
        return {}
    return payload


def _index_previous_results(
    payload: Dict[str, Any],
) -> Dict[Tuple[str, str], Dict[str, Any]]:
    """Index previous results by ``(model_id, label)`` for quick lookup."""
    out: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for row in payload.get("results", []) or []:
        if not isinstance(row, dict):
            continue
        key = (str(row.get("model_id", "")), str(row.get("label", "")))
        out[key] = row
    return out


def merge_last_working(
    row: Dict[str, Any],
    previous_row: Optional[Dict[str, Any]],
    current_date: str,
    current_commit: str,
) -> Dict[str, Any]:
    """Return ``row`` updated with ``last_working_date`` / ``last_working_commit``.

    * If the current run is working, the "last working" timestamp is updated
      to the current run.
    * Otherwise the previous values are carried over (when present), so that
      the dashboard can show the last known good date for a failing cell.
    """
    if row.get("success") == 1:
        row["last_working_date"] = current_date
        row["last_working_commit"] = current_commit or ""
    elif previous_row is not None:
        prev_date = previous_row.get("last_working_date") or ""
        prev_commit = previous_row.get("last_working_commit") or ""
        if prev_date:
            row["last_working_date"] = prev_date
            row["last_working_commit"] = prev_commit
        else:
            row["last_working_date"] = ""
            row["last_working_commit"] = ""
    else:
        row["last_working_date"] = ""
        row["last_working_commit"] = ""
    return row


def _list_xlsx(folder: Optional[str]) -> set:
    """Return the set of all ``.xlsx`` files found recursively under ``folder``."""
    if not folder or not os.path.isdir(folder):
        return set()
    out = set()
    for root, _dirs, files in os.walk(folder):
        for name in files:
            if name.endswith(".xlsx"):
                out.add(os.path.join(root, name))
    return out


def _read_yobx_export_duration(xlsx_path: str) -> Optional[float]:
    """Return ``stat_time_export_and_post_processing`` from the ``extra`` sheet.

    The yobx exporter saves a companion ``.xlsx`` next to every exported
    ``.onnx`` file. The ``extra`` sheet of that workbook is a small
    key/value table with scalar metrics recorded during the export.
    Returns ``None`` when the workbook, sheet, key or value cannot be
    read (missing ``openpyxl``, corrupted file, ...).
    """
    try:  # pragma: no cover - exercised when openpyxl is available
        from openpyxl import load_workbook
    except Exception:  # noqa: BLE001 - optional dependency
        return None
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001 - never fail the recorder on a bad file
        return None
    try:
        if "extra" not in wb.sheetnames:
            return None
        ws = wb["extra"]
        key = "stat_time_export_and_post_processing"
        # ``extra`` is a two-column key/value table; the first row is a
        # header. We do not assume which column holds the key, so scan both.
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            # Find the key cell and pair it with the first non-key cell.
            try:
                idx = row.index(key)
            except ValueError:
                continue
            for j, cell in enumerate(row):
                if j == idx:
                    continue
                value = _to_float(cell)
                if value is not None:
                    return value
            return None
        return None
    finally:
        try:
            wb.close()
        except Exception:  # noqa: BLE001
            pass


def run_validate_one(
    entry: Dict[str, Any],
    exporter_cfg: Dict[str, str],
    verbose: int = 0,
    dump_folder: Optional[str] = None,
    quiet: bool = True,
) -> Any:
    """Run :func:`yobx.torch.validate.validate_model` for one (model, exporter)."""
    # The ``olive-modelbuilder`` column does not use ``validate_model``
    # either: it shells out to the Olive CLI to run the ``ModelBuilder``
    # pass. See :func:`run_olive_modelbuilder`.
    if exporter_cfg.get("exporter") == "olive-modelbuilder":
        return run_olive_modelbuilder(
            entry,
            exporter_cfg,
            verbose=verbose,
            dump_folder=dump_folder,
            quiet=quiet,
        )

    # Lazy import so ``--help`` works without the heavy ``torch`` stack.
    from yobx.torch.validate import validate_model

    kwargs = dict(
        model_id=entry["model"],
        exporter=exporter_cfg["exporter"],
        optimization=exporter_cfg["optimization"],
        dtype=entry.get("dtype", DEFAULT_DTYPE),
        device=entry.get("device", DEFAULT_DEVICE),
        do_run=True,
        quiet=quiet,
        verbose=verbose,
        patch="transformers",
        dump_folder=dump_folder,
        config_overrides={"num_hidden_layers": 2},
        random_weights=True,
    )
    patch_ctx = nullcontext()
    if not entry.get("tokenizer_use_fast", True):
        from transformers import AutoTokenizer

        original_from_pretrained = AutoTokenizer.from_pretrained

        def _force_slow_tokenizer_from_pretrained(
            pretrained_model_name_or_path, *inputs, **tokenizer_kwargs
        ):
            tokenizer_kwargs.setdefault("use_fast", False)
            return original_from_pretrained(
                pretrained_model_name_or_path, *inputs, **tokenizer_kwargs
            )

        patch_ctx = patch.object(
            AutoTokenizer,
            "from_pretrained",
            new=_force_slow_tokenizer_from_pretrained,
        )

    with patch_ctx:
        summary, _data = validate_model(**kwargs)
    return summary


def _olive_precision_for_dtype(dtype: Optional[str]) -> str:
    """Return the Olive ``--precision`` value corresponding to *dtype*."""
    text = (dtype or "").lower()
    if text in ("float16", "fp16", "half"):
        return "fp16"
    if text in ("bfloat16", "bf16"):
        return "bf16"
    if text == "int4":
        return "int4"
    return "fp32"


def run_olive_modelbuilder(
    entry: Dict[str, Any],
    exporter_cfg: Dict[str, str],
    verbose: int = 0,
    dump_folder: Optional[str] = None,
    quiet: bool = True,
) -> Dict[str, Any]:
    """Export *entry* to ONNX using the development version of Olive.

    This shells out to ``olive capture-onnx-graph --use_model_builder``
    twice. Both invocations pass ``--test`` so Olive auto-injects the
    ``OnnxDiscrepancyCheck`` pass into the workflow:

    1. The first command adds ``--dry_run`` so Olive only generates and
       saves the workflow ``config.json`` (already containing the
       ``OnnxDiscrepancyCheck`` pass) without running it.
    2. The second command drops ``--dry_run`` and actually executes the
       same workflow, producing the ONNX graph and the
       ``discrepancy_check_results.json`` metrics file.

    ``--test`` also makes Olive build a randomly initialised,
    two-hidden-layer copy of the HuggingFace architecture and use it as
    the PyTorch reference, so the comparison is between the ONNX graph
    produced by ``ModelBuilder`` and the matching PyTorch test model
    rather than the (potentially gigabytes-sized) full checkpoint.

    The function returns a plain dict that follows the same conventions
    as :func:`yobx.torch.validate.validate_model`'s ``ValidateSummary``
    so that :func:`_normalise_result` can consume it without changes.
    """
    import json
    import os
    import shutil
    import subprocess
    import sys
    import tempfile
    from collections import Counter

    model_id = entry["model"]
    precision = _olive_precision_for_dtype(entry.get("dtype"))

    summary: Dict[str, Any] = {
        "model_id": model_id,
        "export": "",
    }

    own_tmp: Optional[str] = None
    if dump_folder is not None:
        os.makedirs(dump_folder, exist_ok=True)
        output_path = os.path.join(
            dump_folder, f"{model_id.replace('/', '-')}.olive-modelbuilder"
        )
        os.makedirs(output_path, exist_ok=True)
    else:
        own_tmp = tempfile.mkdtemp(prefix="olive_mb_")
        output_path = own_tmp

    base_cmd = [
        sys.executable,
        "-m",
        "olive.cli.launcher",
        "capture-onnx-graph",
        "--model_name_or_path",
        model_id,
        "--output_path",
        output_path,
        "--use_model_builder",
        "--precision",
        precision,
        # ``--test`` makes Olive build a randomly-initialised copy of the
        # HuggingFace architecture with 2 hidden layers and use it as the
        # PyTorch reference. It also auto-injects the
        # ``OnnxDiscrepancyCheck`` pass and dumps the metrics to
        # ``discrepancy_check_results.json`` in the output directory.
        "--test",
    ]
    # First run with ``--dry_run`` so Olive generates ``config.json`` in
    # the output directory (which already includes the
    # ``OnnxDiscrepancyCheck`` pass thanks to ``--test``) without
    # actually executing the workflow. The second run drops
    # ``--dry_run`` and actually runs the same workflow, including the
    # ``OnnxDiscrepancyCheck`` pass.
    dry_run_cmd = [*base_cmd, "--dry_run"]
    cmd = base_cmd

    def _run(cmd_to_run: List[str]) -> Tuple[Optional[Any], Optional[Dict[str, Any]]]:
        try:
            return (
                subprocess.run(
                    cmd_to_run,
                    check=False,
                    capture_output=True,
                    text=True,
                    timeout=30 * 60,
                ),
                None,
            )
        except FileNotFoundError as exc:
            err = {
                "export": "FAILED",
                "error_export": (
                    f"olive CLI is not available: {type(exc).__name__}: {exc}"
                ),
            }
            return None, err
        except subprocess.TimeoutExpired as exc:
            err = {
                "export": "FAILED",
                "error_export": f"olive capture-onnx-graph timed out: {exc}",
            }
            return None, err

    proc, err = _run(dry_run_cmd)
    if err is not None:
        summary.update(err)
        if own_tmp is not None:
            shutil.rmtree(own_tmp, ignore_errors=True)
        return summary
    if proc.returncode != 0:
        stderr = (proc.stderr or "").strip()
        stdout = (proc.stdout or "").strip()
        msg = stderr or stdout or f"olive exited with code {proc.returncode}"
        summary["export"] = "FAILED"
        summary["error_export"] = f"--dry_run failed: {msg}"
        if own_tmp is not None:
            shutil.rmtree(own_tmp, ignore_errors=True)
        return summary

    proc, err = _run(cmd)
    if err is not None:
        summary.update(err)
        if own_tmp is not None:
            shutil.rmtree(own_tmp, ignore_errors=True)
        return summary
    if proc.returncode != 0:
        stderr = (proc.stderr or "").strip()
        stdout = (proc.stdout or "").strip()
        msg = stderr or stdout or f"olive exited with code {proc.returncode}"
        summary["export"] = "FAILED"
        summary["error_export"] = msg
        if own_tmp is not None:
            shutil.rmtree(own_tmp, ignore_errors=True)
        return summary

    # Olive writes the ONNX model and the GenAI config files into
    # ``output_path``. Locate the produced ``.onnx`` (recursively, since
    # Olive may nest it inside a workflow subdirectory) so we can compute
    # node statistics.
    onnx_files: List[str] = []
    for root, _dirs, files in os.walk(output_path):
        for name in files:
            if name.endswith(".onnx"):
                onnx_files.append(os.path.join(root, name))
    if not onnx_files:
        summary["export"] = "FAILED"
        summary["error_export"] = (
            "olive capture-onnx-graph reported success but no .onnx file "
            f"was found under {output_path}."
        )
        if own_tmp is not None:
            shutil.rmtree(own_tmp, ignore_errors=True)
        return summary

    summary["export"] = "OK"
    # ``--test`` makes Olive write ``discrepancy_check_results.json`` in
    # the output directory once the OnnxDiscrepancyCheck pass has run.
    # Surface those metrics on the summary so the dashboard can display
    # the same discrepancy fields as the other columns.
    disc_path: Optional[str] = None
    for root, _dirs, files in os.walk(output_path):
        if "discrepancy_check_results.json" in files:
            disc_path = os.path.join(root, "discrepancy_check_results.json")
            break
    if disc_path is not None:
        try:
            with open(disc_path, encoding="utf-8") as f:
                disc = json.load(f)
        except Exception as exc:  # noqa: BLE001 - best effort, never fail
            summary["discrepancies"] = "FAILED"
            summary["error_discrepancies"] = (
                f"could not read {disc_path}: {type(exc).__name__}: {exc}"
            )
        else:
            status = str(disc.get("status") or "").lower()
            summary["discrepancies"] = "OK" if status == "passed" else "FAILED"
            total = disc.get("total_elements")
            above_atol = disc.get("elements_above_0_01")
            if isinstance(total, int):
                summary["discrepancies_total"] = total
                if isinstance(above_atol, int):
                    summary["discrepancies_ok"] = max(total - above_atol, 0)
            max_abs = disc.get("max_abs_error")
            if isinstance(max_abs, (int, float)):
                summary["discrepancies_max_abs"] = float(max_abs)
            # ``OnnxDiscrepancyCheck`` reports counts at fixed thresholds
            # of 0.01 and 0.1; the 0.01 bucket matches the other columns
            # which use ``atol=0.01`` by default.
            summary["discrepancies_atol"] = 0.01
            failures = disc.get("failures")
            if status != "passed" and failures:
                summary["error_discrepancies"] = "; ".join(str(f) for f in failures)
    else:
        summary["discrepancies"] = "FAILED"
        summary["error_discrepancies"] = (
            "olive capture-onnx-graph --test reported success but "
            f"discrepancy_check_results.json was not found under {output_path}."
        )
    # Pick the largest ONNX file as the main model (external data files
    # may exist alongside but the model file itself is usually the
    # biggest one ending in ``.onnx``).
    main_onnx = max(onnx_files, key=lambda p: os.path.getsize(p))
    try:
        import onnx

        onx = onnx.load(main_onnx, load_external_data=False)
        counts = Counter(n.op_type for n in onx.graph.node)
        summary["n_nodes"] = sum(counts.values())
        top = counts.most_common(5)
        summary["top_op_types"] = ",".join(f"{op}:{cnt}" for op, cnt in top)
    except Exception:  # noqa: BLE001 - statistics are best-effort
        pass

    if own_tmp is not None:
        shutil.rmtree(own_tmp, ignore_errors=True)
    return summary


def detect_task(model_id: str) -> str:
    """Return the HuggingFace task detected for ``model_id``.

    Uses :func:`yobx.torch.validate._detect_task` against the model's
    ``AutoConfig`` (e.g. ``"text-generation"``, ``"fill-mask"``,
    ``"image-classification"``, ``"feature-extraction"``).  Returns an
    empty string when the task cannot be detected (missing dependencies,
    network failure, ...).
    """
    try:  # pragma: no cover - exercised via the recording script
        from transformers import AutoConfig
        from yobx.torch.validate import _detect_task

        config = AutoConfig.from_pretrained(model_id, trust_remote_code=True)
        return _detect_task(config) or ""
    except Exception as exc:  # noqa: BLE001 - never fail the recorder for this
        _log(f"  -> could not detect task for {model_id}: {type(exc).__name__}: {exc}")
        if _is_rate_limit_error(exc):
            # HuggingFace rate-limited us (HTTP 429). Continuing would only
            # make things worse, so abort the whole run instead of silently
            # returning an empty task.
            raise
        return ""


def run_all(
    models: Tuple[Dict[str, Any], ...],
    exporters: Tuple[Dict[str, str], ...],
    limit: Optional[int] = None,
    verbose: int = 0,
    dump_folder: Optional[str] = None,
    quiet: bool = True,
) -> List[Tuple[str, Dict[str, str], Any, float]]:
    """Run ``validate_model`` for every (model, exporter) combination."""
    import shutil
    import tempfile

    items = list(models)
    if limit is not None:
        items = items[:limit]
    out: List[Tuple[str, Dict[str, str], Any, float]] = []
    # Cache of synthetic "HuggingFace Hub access" summaries keyed by model id.
    # Populated the first time a model fails at the ``error_config`` step with
    # an unreachable-Hub / gated-model error so the remaining exporters for
    # that same model can be short-circuited (they would all fail at the same
    # config-loading step with the exact same message).
    hub_access_failures: Dict[str, Dict[str, Any]] = {}
    for entry in items:
        model_id = entry["model"]
        for exporter_cfg in exporters:
            _log(
                f"Validating {model_id} (dtype={entry.get('dtype')}, "
                f"device={entry.get('device')}) with "
                f"exporter={exporter_cfg['exporter']} "
                f"optimization={exporter_cfg['optimization']}..."
            )
            cached = hub_access_failures.get(model_id)
            if cached is not None:
                _log(
                    "  -> skipping: previous exporter already failed to "
                    "reach the HuggingFace Hub for this model."
                )
                out.append((model_id, exporter_cfg, dict(cached), 0.0))
                continue
            # For yobx, the export time we want to record is the
            # ``stat_time_export_and_post_processing`` metric stored in
            # the ``extra`` sheet of the workbook saved next to the
            # ONNX file. We therefore make sure that a dump folder is
            # always passed for yobx runs, even when the user did not
            # supply one (in which case we use a per-cell temporary
            # directory which is removed once the metric is read).
            is_yobx = exporter_cfg["exporter"] == "yobx"
            tmp_dump: Optional[str] = None
            effective_dump = dump_folder
            if is_yobx and effective_dump is None:
                tmp_dump = tempfile.mkdtemp(prefix="yobx_dump_")
                effective_dump = tmp_dump
            # Snapshot the xlsx files present before the export so we can
            # detect the workbook(s) the yobx exporter just produced.
            pre_xlsx = _list_xlsx(effective_dump) if is_yobx else set()
            start = time.monotonic()
            try:
                summary = run_validate_one(
                    entry,
                    exporter_cfg,
                    verbose=verbose,
                    dump_folder=effective_dump,
                    quiet=quiet,
                )
            except Exception as exc:  # noqa: BLE001 - we never want to crash CI
                _log(f"  -> raised: {type(exc).__name__}: {exc}")
                if _is_rate_limit_error(exc):
                    # HuggingFace rate-limited us (HTTP 429). Continuing
                    # would only make things worse, so abort the whole run
                    # and let the workflow surface the failure.
                    _log(
                        "  -> aborting: HuggingFace returned HTTP 429 "
                        "(rate limit). Stopping the recording action."
                    )
                    raise
                # Build a synthetic summary so the dashboard can still display
                # a failed cell with a meaningful error message.
                summary = {
                    "model_id": model_id,
                    "export": "FAILED",
                    "error_export": f"{type(exc).__name__}: {exc}",
                }
                if not quiet:
                    raise
            duration_s = time.monotonic() - start
            # For yobx, prefer the ``stat_time_export_and_post_processing``
            # metric recorded in the ``extra`` sheet of the generated
            # workbook over the wall-clock time, which also includes the
            # discrepancy check and other unrelated work.
            if is_yobx and effective_dump:
                new_xlsx = _list_xlsx(effective_dump) - pre_xlsx

                def _safe_mtime(p: str) -> float:
                    try:
                        return os.path.getmtime(p)
                    except OSError:
                        return 0.0

                for xlsx_path in sorted(new_xlsx, key=_safe_mtime, reverse=True):
                    metric = _read_yobx_export_duration(xlsx_path)
                    if metric is not None:
                        duration_s = metric
                        break
            if tmp_dump is not None:
                shutil.rmtree(tmp_dump, ignore_errors=True)
            _log(f"  -> done in {duration_s:.2f}s")
            out.append((model_id, exporter_cfg, summary, duration_s))
            # If this cell failed at the ``error_config`` step because the
            # HuggingFace Hub could not be reached (real network outage, or
            # — more commonly in CI — a gated model that the available
            # token cannot download), reuse the same synthetic summary for
            # the remaining exporters of the same model instead of hitting
            # the Hub three more times.
            if model_id not in hub_access_failures:
                step, message = _first_error(summary)
                if step == "config" and _is_hf_hub_access_error(message):
                    hub_access_failures[model_id] = {
                        "model_id": model_id,
                        "export": "FAILED",
                        "error_config": message,
                    }
    return out


def build_payload(
    raw_results: List[Tuple[str, Dict[str, str], Any, float]],
    models: Tuple[Dict[str, Any], ...],
    exporters: Tuple[Dict[str, str], ...],
    dtype: str,
    device: str,
    commit: Optional[str],
    previous_payload: Optional[Dict[str, Any]] = None,
    now: Optional[str] = None,
    tasks: Optional[Dict[str, str]] = None,
) -> Dict[str, Any]:
    """Assemble the JSON snapshot to be written to disk."""
    current_date = now or dt.datetime.now(tz=dt.timezone.utc).strftime(
        "%Y-%m-%dT%H:%M:%SZ"
    )
    previous_index = _index_previous_results(previous_payload or {})
    previous_tasks = (previous_payload or {}).get("tasks") or {}
    # Map model id -> entry dict for per-model lookups.
    entries_by_id: Dict[str, Dict[str, Any]] = {e["model"]: e for e in models}
    model_ids: List[str] = [e["model"] for e in models]
    resolved_tasks: Dict[str, str] = {}
    for m in model_ids:
        entry = entries_by_id.get(m, {})
        if tasks and tasks.get(m):
            resolved_tasks[m] = tasks[m]
        elif entry.get("task"):
            resolved_tasks[m] = entry["task"]
        elif previous_tasks.get(m):
            resolved_tasks[m] = previous_tasks[m]
        else:
            resolved_tasks[m] = ""

    results: List[Dict[str, Any]] = []
    totals: Dict[str, Dict[str, int]] = {}
    for label in (e["label"] for e in exporters):
        totals[label] = {"success": 0, "failure": 0, "total": 0}

    for item in raw_results:
        if len(item) == 4:
            model_id, exporter_cfg, summary, duration_s = item
        else:
            model_id, exporter_cfg, summary = item
            duration_s = None
        row = _normalise_result(
            model_id,
            exporter_cfg,
            summary,
            duration_s,
            model_atol=_to_float(entries_by_id.get(model_id, {}).get("atol")),
        )
        row["task"] = resolved_tasks.get(model_id, "")
        entry = entries_by_id.get(model_id, {})
        row["dtype"] = entry.get("dtype", dtype)
        row["device"] = entry.get("device", device)
        row["atol"] = _to_float(entry.get("atol"))
        previous_row = previous_index.get((model_id, exporter_cfg["label"]))
        merge_last_working(row, previous_row, current_date, commit or "")
        results.append(row)
        bucket = totals.setdefault(
            exporter_cfg["label"], {"success": 0, "failure": 0, "total": 0}
        )
        bucket["total"] += 1
        if row["success"] == 1:
            bucket["success"] += 1
        else:
            bucket["failure"] += 1

    return {
        "date": current_date,
        "commit": commit or "",
        "versions": collect_versions(),
        "dtype": dtype,
        "device": device,
        "models": model_ids,
        "model_entries": [dict(e) for e in models],
        "tasks": resolved_tasks,
        "exporters": [dict(e) for e in exporters],
        "totals": totals,
        "results": results,
    }


def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--cache-dir",
        default=os.path.join("cache_data"),
        help="Root directory of the JSON cache (default: %(default)s).",
    )
    parser.add_argument(
        "--repo",
        default="yet-another-onnx-builder",
        help="Sub-directory of the cache to write into (default: %(default)s).",
    )
    parser.add_argument(
        "--model",
        action="append",
        dest="models",
        help=(
            "HuggingFace model id to validate. May be repeated. "
            f"Defaults to: {', '.join(e['model'] for e in DEFAULT_MODELS)}."
        ),
    )
    parser.add_argument(
        "--dtype",
        default=DEFAULT_DTYPE,
        help=f"Dtype passed to validate_model (default: {DEFAULT_DTYPE}).",
    )
    parser.add_argument(
        "--device",
        default=DEFAULT_DEVICE,
        help=f"Device passed to validate_model (default: {DEFAULT_DEVICE}).",
    )
    parser.add_argument(
        "--commit",
        default=os.environ.get("YOBX_COMMIT", ""),
        help="Commit SHA of yet-another-onnx-builder used for the snapshot.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Limit to the first N models (mainly for local testing).",
    )
    parser.add_argument(
        "--test",
        action="store_true",
        help=(
            "Smoke-test mode: ignore the default model list and run only "
            "against a single tiny test model (``arnir0/Tiny-LLM``). This "
            "is intended to quickly validate that a newly-added runtime "
            "(such as ``olive-modelbuilder``) is wired correctly without "
            "exercising the full set of default models."
        ),
    )
    parser.add_argument(
        "--verbose",
        type=int,
        default=0,
        help="Verbosity level forwarded to validate_model (default: 0).",
    )
    parser.add_argument(
        "--quiet",
        dest="quiet",
        action="store_true",
        default=True,
        help=(
            "Forward quiet=True to validate_model so per-model output is "
            "suppressed (default)."
        ),
    )
    parser.add_argument(
        "--no-quiet",
        dest="quiet",
        action="store_false",
        help=(
            "Forward quiet=False to validate_model so the underlying "
            "tracebacks and progress output are shown."
        ),
    )
    parser.add_argument(
        "--dump-folder",
        default=None,
        help=(
            "Folder where validate_model dumps its intermediate artefacts "
            "(ONNX files, captured inputs, ...). Intended for local runs. "
            "The folder is created if missing and the script changes its "
            "working directory to it before running so any relative paths "
            "(including --cache-dir) are resolved inside it."
        ),
    )
    return parser.parse_args(argv)


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv)
    if args.test:
        # ``--test`` is a smoke-test shortcut: ignore both the default
        # model list and any ``--model`` overrides, and run against a
        # single tiny model so a newly-added runtime can be validated
        # quickly. ``arnir0/Tiny-LLM`` is a small Llama-architecture
        # checkpoint that is also supported by Olive's ``ModelBuilder``.
        models = (
            _coerce_model_entry(
                "arnir0/Tiny-LLM",
                default_dtype=args.dtype,
                default_device=args.device,
            ),
        )
    elif args.models:
        models = tuple(
            _coerce_model_entry(m, default_dtype=args.dtype, default_device=args.device)
            for m in args.models
        )
    else:
        models = tuple(
            _coerce_model_entry(e, default_dtype=args.dtype, default_device=args.device)
            for e in DEFAULT_MODELS
        )
    exporters = DEFAULT_EXPORTERS

    dump_folder: Optional[str] = None
    if args.dump_folder:
        dump_folder = os.path.abspath(args.dump_folder)
        os.makedirs(dump_folder, exist_ok=True)
        _log(f"Using dump folder: {dump_folder} (chdir into it)")
        os.chdir(dump_folder)

    out_dir = os.path.join(args.cache_dir, args.repo)
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "model_validate.json")

    previous_payload = _load_existing_cache(out_path)

    tasks = {e["model"]: e.get("task") or detect_task(e["model"]) for e in models}

    raw_results = run_all(
        models=models,
        exporters=exporters,
        limit=args.limit,
        verbose=args.verbose,
        dump_folder=dump_folder,
        quiet=args.quiet,
    )
    payload = build_payload(
        raw_results=raw_results,
        models=models,
        exporters=exporters,
        dtype=args.dtype,
        device=args.device,
        commit=args.commit,
        previous_payload=previous_payload,
        tasks=tasks,
    )

    with open(out_path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2, sort_keys=True, allow_nan=False)
        fh.write("\n")

    _log(
        f"Wrote {out_path}: {len(payload['results'])} results, "
        f"{len(models)} models, {len(exporters)} exporters."
    )
    for label, bucket in payload["totals"].items():
        _log(
            f"  {label}: {bucket['success']}/{bucket['total']} working "
            f"({bucket['failure']} failed)."
        )
    return 0


if __name__ == "__main__":
    sys.exit(main())
